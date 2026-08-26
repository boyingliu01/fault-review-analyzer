"""Tests for GAP fixes G10-G13 (functional GAP analysis remediation).

Covers:
- G10: RecurrenceDetector detects recurring patterns + RecurrencePattern model
- G11: ReportGenerator Excel export
- G13: get_code_diffs alias + graceful degradation
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock, patch

import pytest

from src.api.client import APIClient
from src.feedback.models import RecurrencePattern
from src.feedback.recurrence_detector import RecurrenceDetector
from src.report.generator import ReportData, ReportFormat, ReportType

# --- G10: RecurrenceDetector ---


class TestRecurrenceDetector:
    def test_detect_returns_recurrence_patterns(self):
        """G10: Similar tasks are grouped into recurrence patterns."""
        detector = RecurrenceDetector(similarity_threshold=0.3)
        tasks = [
            {"task_id": "1", "title": "数据库连接超时", "description": "连接池耗尽导致超时"},
            {"task_id": "2", "title": "数据库连接超时", "description": "连接池耗尽导致超时"},
            {"task_id": "3", "title": "并发问题", "description": "竞态条件"},
        ]

        patterns = detector.detect(tasks)

        # 前两个任务相似，第三个不同 → 应产生 1 个复发模式
        assert len(patterns) == 1
        pattern = patterns[0]
        assert isinstance(pattern, RecurrencePattern)
        assert pattern.occurrence_count == 2
        assert set(pattern.task_ids) == {"1", "2"}

    def test_detect_no_recurrence(self):
        """G10: Distinct tasks produce no recurrence patterns."""
        detector = RecurrenceDetector(similarity_threshold=0.9)
        tasks = [
            {"task_id": "1", "title": "数据库超时"},
            {"task_id": "2", "title": "前端样式错乱"},
        ]

        patterns = detector.detect(tasks)
        assert patterns == []

    def test_detect_empty_tasks(self):
        """G10: Empty task list returns empty patterns."""
        detector = RecurrenceDetector()
        assert detector.detect([]) == []

    def test_detect_single_task_no_pattern(self):
        """G10: Single task does not form a recurrence pattern."""
        detector = RecurrenceDetector(similarity_threshold=0.1)
        tasks = [{"task_id": "1", "title": "唯一故障"}]

        patterns = detector.detect(tasks)
        assert patterns == []

    def test_recurrence_pattern_model_defaults(self):
        """G10: RecurrencePattern has sensible defaults."""
        pattern = RecurrencePattern(name="test", task_ids=["1"])
        assert pattern.occurrence_count == 0
        assert pattern.confidence == 0.0
        assert pattern.similarity_threshold == 0.7
        assert pattern.severity == "medium"

    def test_detect_with_embedding_generator(self):
        """G10: embedding_generator is used for similarity when provided."""
        detector = RecurrenceDetector(similarity_threshold=0.3)
        mock_embedding = MagicMock()
        mock_embedding.embed_batch = AsyncMock(return_value=[[1.0, 0.0], [0.0, 1.0]])
        detector._embedding_generator = mock_embedding

        # 直接测试 _compute_similarity：正交向量相似度为 0
        sim = detector._compute_similarity("任务A", "任务B")
        assert sim == 0.0


# --- G11: Excel export ---


class TestExcelExport:
    def test_report_format_has_excel(self):
        """G11: ReportFormat includes EXCEL."""
        assert ReportFormat.EXCEL.value == "excel"

    def test_get_format_extension_excel(self):
        """G11: Excel extension is xlsx."""
        from src.report.generator import ReportGenerator

        gen = ReportGenerator()
        assert gen._get_format_extension(ReportFormat.EXCEL) == "xlsx"

    def test_generate_excel_creates_file(self, tmp_path):
        """G11: generate() writes an Excel file when format is EXCEL."""
        from src.report.generator import ReportGenerator

        data = ReportData(
            title="测试报告",
            type=ReportType.SUMMARY,
            generated_at=datetime.now(),
            summary={"total_tasks": 10, "cluster_count": 3},
            tables=[],
        )

        gen = ReportGenerator(output_dir=tmp_path)
        result = gen.generate(data, ReportFormat.EXCEL, filename="report")

        assert isinstance(result, Path)
        assert result.exists()
        assert result.suffix == ".xlsx"

        # 验证文件可被 openpyxl 读取
        import openpyxl

        wb = openpyxl.load_workbook(result)
        assert "摘要" in wb.sheetnames

    def test_generate_excel_requires_filename(self, tmp_path):
        """G11: Excel export requires filename and output_dir."""
        from src.report.generator import ReportGenerator

        data = ReportData(
            title="测试",
            type=ReportType.SUMMARY,
            generated_at=datetime.now(),
            summary={},
        )
        gen = ReportGenerator(output_dir=tmp_path)

        with pytest.raises(ValueError, match="requires both filename and output_dir"):
            gen.generate(data, ReportFormat.EXCEL, filename=None)

    def test_generate_excel_with_tables(self, tmp_path):
        """G11: Excel export includes table sheets."""
        from src.report.generator import ReportGenerator, TableData

        data = ReportData(
            title="测试报告",
            type=ReportType.DETAILED,
            generated_at=datetime.now(),
            summary={"total": 5},
            tables=[
                TableData(
                    title="违规统计",
                    headers=["rule_id", "count"],
                    rows=[["J000001", 3], ["J000002", 2]],
                )
            ],
        )

        gen = ReportGenerator(output_dir=tmp_path)
        result = gen.generate(data, ReportFormat.EXCEL, filename="report")

        import openpyxl

        wb = openpyxl.load_workbook(result)
        assert "表格数据" in wb.sheetnames


# --- G13: get_code_diffs alias + graceful degradation ---


class TestGetCodeDiffs:
    @pytest.mark.asyncio
    async def test_get_code_diffs_aliases_get_commits(self):
        """G13: get_code_diffs calls get_commits with with_content=True."""
        client = APIClient(base_url="https://api.example.com")
        mock_commits = [MagicMock()]

        with patch.object(
            client, "get_commits", new_callable=AsyncMock, return_value=mock_commits
        ) as mock_get:
            result = await client.get_code_diffs(12345)

        assert result == mock_commits
        mock_get.assert_called_once_with(12345, with_content=True)

    @pytest.mark.asyncio
    async def test_get_code_diffs_degrades_on_connection_error(self):
        """G13: get_code_diffs returns [] on connection error (graceful)."""
        from src.api.exceptions import APIConnectionError

        client = APIClient(base_url="https://api.example.com")

        with patch.object(
            client,
            "get_commits",
            new_callable=AsyncMock,
            side_effect=APIConnectionError("connection refused"),
        ):
            result = await client.get_code_diffs(12345)

        assert result == []

    @pytest.mark.asyncio
    async def test_get_full_task_degrades_on_commits_error(self):
        """G13: get_full_task treats commit-fetch failure as no code changes."""
        from src.api.exceptions import APIConnectionError

        client = APIClient(base_url="https://api.example.com")
        mock_task = MagicMock()
        mock_task.task_id = 12345
        mock_task.development = None

        with (
            patch.object(client, "get_task", new_callable=AsyncMock, return_value=mock_task),
            patch.object(
                client,
                "get_commits",
                new_callable=AsyncMock,
                side_effect=APIConnectionError("down"),
            ),
            patch.object(
                client,
                "get_production_info",
                new_callable=AsyncMock,
                return_value=MagicMock(),
            ),
            patch.object(
                client,
                "get_fault_analysis",
                new_callable=AsyncMock,
                return_value={},
            ),
        ):
            result = await client.get_full_task(12345)

        # development 保持 None（优雅降级，不抛异常）
        assert result.development is None
